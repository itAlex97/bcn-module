import os
import sys
from copy import copy

from PySide6 import QtCore, QtWidgets, QtGui

from export import guardar_resultado
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import tempfile
from main import DEFAULT_COMPONENTS_PATH, calcular_resultado, obtener_partes_afectadas


class BomWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("BOM Comparison Tool")
        self.resize(920, 520)

        # Se guardan despues de procesar para poder exportar/abrir el resultado.
        self.result_df = None
        self.result_path = None

        self._setup_ui()
        self._apply_stylesheet()

    def _setup_ui(self):
        """Arma la ventana principal y sus tres zonas: entrada, acciones y log."""
        main_widget = QtWidgets.QWidget()
        main_layout = QtWidgets.QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(12)

        title = QtWidgets.QLabel("BOM Comparison")
        title_font = title.font()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        main_layout.addWidget(title)

        main_layout.addWidget(self._create_input_section())
        main_layout.addWidget(self._create_action_section())
        main_layout.addWidget(self._create_results_section(), 1)

        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)

    def _create_input_section(self):
        """Crea los campos para seleccionar Charted y MFGPro."""
        container = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        layout.addWidget(QtWidgets.QLabel("Charted BOM"))

        self.charted_edit = QtWidgets.QLineEdit()
        self.charted_edit.setPlaceholderText("Charted BOM")
        self.charted_edit.setMinimumHeight(34)
        charted_row = QtWidgets.QHBoxLayout()
        charted_row.addWidget(self.charted_edit, 1)
        charted_btn = self._create_button("Buscar", self.select_charted, style="secondary")
        charted_row.addWidget(charted_btn)
        layout.addLayout(charted_row)

        layout.addWidget(QtWidgets.QLabel("MFGPro"))

        self.mfgpro_edit = QtWidgets.QLineEdit()
        self.mfgpro_edit.setPlaceholderText("MFGPro")
        self.mfgpro_edit.setMinimumHeight(34)
        mfgpro_row = QtWidgets.QHBoxLayout()
        mfgpro_row.addWidget(self.mfgpro_edit, 1)
        mfgpro_btn = self._create_button("Buscar", self.select_mfgpro, style="secondary")
        mfgpro_row.addWidget(mfgpro_btn)
        layout.addLayout(mfgpro_row)

        container.setLayout(layout)
        return container

    def _create_action_section(self):
        """Crea botones de proceso/exportacion y el texto de estado."""
        layout = QtWidgets.QHBoxLayout()
        layout.setSpacing(8)

        self.process_btn = self._create_button("Comparar", self.run_process, style="primary")
        self.process_btn.setMinimumHeight(36)
        self.process_btn.setMinimumWidth(110)

        self.save_btn = self._create_button(
            "Guardar resultado", self.save_result, style="success"
        )
        self.save_btn.setEnabled(False)
        self.save_btn.setMinimumHeight(36)
        self.save_btn.setMinimumWidth(140)

        self.open_btn = self._create_button(
            "Abrir resultado", self.open_result, style="secondary"
        )
        self.open_btn.setEnabled(False)
        self.open_btn.setMinimumHeight(36)
        self.open_btn.setMinimumWidth(130)

        self.export_bcn_btn = self._create_button(
            "Exportar BCN", self.export_bcn, style="success"
        )
        self.export_bcn_btn.setEnabled(False)
        self.export_bcn_btn.setMinimumHeight(36)
        self.export_bcn_btn.setMinimumWidth(140)

        layout.addWidget(self.process_btn)
        layout.addWidget(self.save_btn)
        layout.addWidget(self.open_btn)
        layout.addWidget(self.export_bcn_btn)
        layout.addStretch()

        self.status_label = QtWidgets.QLabel("Listo")
        status_font = self.status_label.font()
        status_font.setPointSize(9)
        self.status_label.setFont(status_font)
        layout.addWidget(self.status_label, 0, QtCore.Qt.AlignmentFlag.AlignVCenter)

        container = QtWidgets.QWidget()
        container.setLayout(layout)
        return container

    def _create_results_section(self):
        """Crea el panel de mensajes para que el usuario vea el avance."""
        container = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)

        self.log = QtWidgets.QPlainTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

        container.setLayout(layout)
        return container

    def _create_button(self, text, callback, style="primary"):
        """Crea botones con estilo consistente y conecta su accion."""
        btn = QtWidgets.QPushButton(text)
        btn.clicked.connect(callback)
        btn.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))

        btn.setObjectName(style)
        return btn

    def _apply_stylesheet(self):
        """Centraliza colores y bordes de la interfaz."""
        self.setStyleSheet(
            "QMainWindow { background-color: #F6F7FB; }"
            "QLabel { color: #1F2937; }"
            "QLineEdit { background: #FFFFFF; color: #111827; border: 1px solid #D6DAE3; "
            "border-radius: 6px; padding: 6px 8px; }"
            "QLineEdit::placeholder { color: #9CA3AF; }"
            "QPlainTextEdit { background: #FFFFFF; color: #111827; border: 1px solid #E1E4EA; "
            "border-radius: 6px; padding: 8px; font-family: Consolas; font-size: 10px; }"
            "QPushButton { background: #FFFFFF; border: 1px solid #D6DAE3; "
            "border-radius: 6px; padding: 6px 12px; color: #1F2937; }"
            "QPushButton:hover { background: #EEF2F7; }"
            "QPushButton:disabled { color: #9CA3AF; background: #F3F4F6; }"
            "QPushButton#primary { background: #2563EB; color: #FFFFFF; border: none; }"
            "QPushButton#primary:hover { background: #1D4ED8; }"
            "QPushButton#success { background: #16A34A; color: #FFFFFF; border: none; }"
            "QPushButton#success:hover { background: #15803D; }"
            "QMessageBox { background-color: #FFFFFF; color: #111827; }"
            "QMessageBox QLabel { color: #111827; }"
            "QMessageBox QPushButton { background: #FFFFFF; border: 1px solid #D6DAE3; "
            "border-radius: 6px; padding: 6px 12px; color: #111827; }"
            "QMessageBox QPushButton:hover { background: #EEF2F7; }"
        )

    def select_charted(self):
        """Selecciona el archivo original de Charted BOM."""
        path = self._select_excel_file("Selecciona el Charted BOM")
        if path:
            self.charted_edit.setText(path)

    def select_mfgpro(self):
        """Selecciona el archivo original de MFGPro."""
        path = self._select_excel_file("Selecciona el MFGPro")
        if path:
            self.mfgpro_edit.setText(path)

    def _select_excel_file(self, title):
        """Abre un dialogo reutilizable para elegir archivos Excel."""
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            title,
            "",
            "Excel (*.xlsx *.xls);;Todos (*.*)",
        )
        return path

    def run_process(self):
        """Valida entradas, ejecuta la comparacion y deja listo el resultado."""
        charted_path = self.charted_edit.text().strip()
        mfgpro_path = self.mfgpro_edit.text().strip()

        if not charted_path:
            QtWidgets.QMessageBox.warning(self, "Falta Charted", "Selecciona el Charted BOM.")
            return
        if not mfgpro_path:
            QtWidgets.QMessageBox.warning(self, "Falta MFGPro", "Selecciona el MFGPro.")
            return

        try:
            self.status_label.setText("Procesando...")
            QtWidgets.QApplication.processEvents()
            self.log_message("Cargando datos...")
            (
                df_resultado,
                charted_partes,
                charted_modelos,
                mfg_partes,
                mfg_modelos,
                resumen,
            ) = calcular_resultado(
                charted_path,
                ruta_mfgpro=mfgpro_path,
            )
            self._log_result_summary(
                df_resultado,
                charted_partes,
                charted_modelos,
                mfg_partes,
                mfg_modelos,
                resumen,
            )
            self.result_df = df_resultado
            self.save_btn.setEnabled(True)
            self.export_bcn_btn.setEnabled(True)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Error", str(exc))
            self.log_message(f"BD componentes esperada: {DEFAULT_COMPONENTS_PATH}")
            self.status_label.setText("Error")
            return

        self.status_label.setText("Completado")
        QtWidgets.QMessageBox.information(self, "Listo", "Comparacion finalizada.")

    def save_result(self):
        """Guarda el DataFrame de diferencias en un archivo Excel."""
        if self.result_df is None:
            QtWidgets.QMessageBox.warning(self, "Sin resultado", "Primero ejecuta el proceso.")
            return

        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Guardar resultado",
            "Resultado_BOM.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return

        try:
            guardar_resultado(self.result_df, path)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Error", str(exc))
            return

        self.result_path = path
        self.open_btn.setEnabled(True)
        self.export_bcn_btn.setEnabled(True)
        self.log_message(f"Resultado guardado: {path}")

    def open_result(self):
        """Abre el ultimo archivo guardado desde Windows."""
        if not self.result_path or not os.path.exists(self.result_path):
            QtWidgets.QMessageBox.warning(self, "No encontrado", "No existe el archivo guardado.")
            return
        os.startfile(self.result_path)

    def _log_result_summary(
        self,
        df_resultado,
        charted_partes,
        charted_modelos,
        mfg_partes,
        mfg_modelos,
        resumen,
    ):
        """Muestra en el log los conteos principales del proceso."""
        self.log_message(
            f"Charted: {charted_partes} numeros de parte, {charted_modelos} niveles"
        )
        self.log_message(
            f"MFGPro:  {mfg_partes} numeros de parte, {mfg_modelos} niveles"
        )
        if df_resultado.empty:
            self.log_message("No se encontraron diferencias entre los BOMs.")
        else:
            self.log_message(f"Diferencias: {len(df_resultado)} -> {resumen}")
            self.log_message("Numeros de parte afectados:")
            for parte in obtener_partes_afectadas(df_resultado):
                self.log_message(f"   - {parte}")

    def export_bcn(self):
        """Crea una copia del template bcn_template, pega los valores del resultado en COMPARE y guarda."""
        if self.result_df is None:
            QtWidgets.QMessageBox.warning(self, "Sin resultado", "Primero ejecuta el proceso.")
            return

        # Preparar origen: usar archivo guardado si existe, si no, generar temporalmente
        temp_created = False
        try:
            if self.result_path and os.path.exists(self.result_path):
                src_path = self.result_path
            else:
                fd, tmp = tempfile.mkstemp(suffix=".xlsx")
                os.close(fd)
                guardar_resultado(self.result_df, tmp)
                src_path = tmp
                temp_created = True

            src_wb = load_workbook(src_path, data_only=True)
            src_ws = src_wb["Resultado"] if "Resultado" in src_wb.sheetnames else src_wb.active

            # Obtener internal number desde el Charted original (celda B4)
            internal = "UNKNOWN"
            try:
                charted_path = self.charted_edit.text().strip()
                if charted_path:
                    ch_wb = load_workbook(charted_path, data_only=True)
                    sheet_name = "Detail" if "Detail" in ch_wb.sheetnames else ch_wb.sheetnames[0]
                    ch_ws = ch_wb[sheet_name]
                    val = ch_ws["B4"].value
                    if val is not None:
                        internal = str(val)
            except Exception:
                pass

            # Cargar template y localizar la hoja COMPARE
            template_path = os.path.join(os.path.dirname(__file__), "data", "bcn_template.xlsx")
            if not os.path.exists(template_path):
                QtWidgets.QMessageBox.warning(self, "Template no encontrado", f"No se encontro: {template_path}")
                return

            tpl_wb = load_workbook(template_path)
            compare_ws = None
            for name in tpl_wb.sheetnames:
                if name.lower() == "compare":
                    compare_ws = tpl_wb[name]
                    break
            if compare_ws is None:
                QtWidgets.QMessageBox.warning(self, "Hoja COMPARE no encontrada", "El template no contiene la hoja COMPARE.")
                return

            if "BCN" not in tpl_wb.sheetnames:
                QtWidgets.QMessageBox.warning(self, "Hoja BCN no encontrada", "El template no contiene la hoja BCN.")
                return
            bcn_ws = tpl_wb["BCN"]

            max_row = src_ws.max_row
            max_col = src_ws.max_column

            # Primero: pegar todos los valores del resultado en la hoja COMPARE
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    compare_ws.cell(row=r, column=c).value = src_ws.cell(row=r, column=c).value

            # Detectar dónde empieza la sección de comentarios para insertar filas sin romperla.
            comments_row = None
            for rr in range(87, bcn_ws.max_row + 1):
                if bcn_ws.cell(row=rr, column=3).value == "COMMENTS":
                    comments_row = rr
                    break
            if comments_row is None:
                comments_row = 87

            display_start = 51
            total_changes = max(0, max_row - 2)
            spacer_rows = 2
            template_slots = comments_row - display_start
            required_slots = total_changes + spacer_rows

            if required_slots > template_slots:
                extra_rows = required_slots - template_slots
                ranges_to_shift = [str(rng) for rng in bcn_ws.merged_cells.ranges if rng.min_row >= comments_row]
                for rng_text in ranges_to_shift:
                    bcn_ws.unmerge_cells(rng_text)
                bcn_ws.insert_rows(comments_row, amount=extra_rows)
                for rng_text in ranges_to_shift:
                    start, end = rng_text.split(":")
                    start_col = "".join(ch for ch in start if ch.isalpha())
                    start_row = int("".join(ch for ch in start if ch.isdigit())) + extra_rows
                    end_col = "".join(ch for ch in end if ch.isalpha())
                    end_row = int("".join(ch for ch in end if ch.isdigit())) + extra_rows
                    bcn_ws.merge_cells(f"{start_col}{start_row}:{end_col}{end_row}")
                comments_row += extra_rows

            # Copiar la fila 51 como plantilla para cada cambio.
            source_row = 51
            source_height = bcn_ws.row_dimensions[source_row].height
            source_merges = [mr for mr in bcn_ws.merged_cells.ranges if mr.min_row == source_row and mr.max_row == source_row]

            def clone_template_row(target_row, copy_values=True, height_override=None):
                """Copia formato/bordes de la fila plantilla a otra fila sin tocar celdas fusionadas."""
                if height_override is not None:
                    bcn_ws.row_dimensions[target_row].height = height_override
                elif source_height is not None:
                    bcn_ws.row_dimensions[target_row].height = source_height

                for cc in range(1, bcn_ws.max_column + 1):
                    src_cell = bcn_ws.cell(row=source_row, column=cc)
                    dst_cell = bcn_ws.cell(row=target_row, column=cc)

                    if isinstance(src_cell, MergedCell) or isinstance(dst_cell, MergedCell):
                        continue

                    dst_cell._style = copy(src_cell._style)
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = copy(src_cell.protection)

                    if not copy_values:
                        dst_cell.value = None
                        continue

                    if cc == 3:
                        dst_cell.value = target_row - display_start + 1
                    elif isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                        dst_cell.value = Translator(src_cell.value, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)
                    else:
                        dst_cell.value = src_cell.value

            def apply_row_rules(target_row):
                """Reaplica las formulas de reglas de la fila 51 a AN/AO/AP en la fila destino."""
                for col_idx in (40, 41, 42):
                    src_cell = bcn_ws.cell(row=source_row, column=col_idx)
                    dst_cell = bcn_ws.cell(row=target_row, column=col_idx)
                    if isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                        dst_cell.value = Translator(src_cell.value, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)
            for index in range(1, total_changes + 1):
                target_row = display_start + index - 1
                if target_row >= comments_row:
                    # Si hiciera falta más espacio por algún motivo, insertar antes de COMMENTS.
                    bcn_ws.insert_rows(comments_row, amount=1)
                    comments_row += 1

                if target_row > 84:
                    # Recrear las fusiones de la fila plantilla para las filas nuevas.
                    for mr in source_merges:
                        shifted = f"{get_column_letter(mr.min_col)}{target_row}:{get_column_letter(mr.max_col)}{target_row}"
                        if shifted not in [str(rng) for rng in bcn_ws.merged_cells.ranges]:
                            bcn_ws.merge_cells(shifted)

                clone_template_row(target_row, copy_values=True)
                apply_row_rules(target_row)

            # Dejar siempre dos filas en blanco al final de la lista, con alturas fijas.
            blank_row_1 = display_start + total_changes
            blank_row_2 = blank_row_1 + 1
            for rr, height in ((blank_row_1, 30), (blank_row_2, 10)):
                clone_template_row(rr, copy_values=False, height_override=height)

            # Preguntar dónde guardar con nombre por defecto
            default_name = f"BCN - {internal}.xlsx"
            path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Guardar BCN", default_name, "Excel (*.xlsx)")
            if not path:
                return

            tpl_wb.save(path)
            QtWidgets.QMessageBox.information(self, "Listo", f"BCN guardado: {path}")

        finally:
            if temp_created:
                try:
                    os.remove(tmp)
                except Exception:
                    pass

        

    def log_message(self, message):
        """Agrega una linea al log visible de la ventana."""
        self.log.appendPlainText(message)


def run_app():
    """Punto de entrada de la aplicacion grafica."""
    app = QtWidgets.QApplication(sys.argv)
    win = BomWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    run_app()
