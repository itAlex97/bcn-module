import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def guardar_resultado(df_result, ruta_salida):
    """Genera el Excel final con formato parecido al reporte esperado."""
    if not df_result.empty:
        df_result = df_result.sort_values(["Model", "Part Nbr."]).reset_index(drop=True)

    # Libro y hoja de salida.
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    # Anchos fijos del formato original.
    column_widths = {
        "A": 15.43,
        "B": 13.86,
        "C": 19.57,
        "D": 13.86,
        "E": 7,
        "F": 26.71,
    }

    # Estilos reutilizables para no crear objetos en cada celda.
    font_normal = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left")
    right = Alignment(horizontal="right")

    # Posiciones de columnas en el Excel final.
    col_nivel = 1
    col_part = 2
    col_qty_new = 3
    col_qty_old = 4
    col_unit = 5
    col_desc = 6
    total_cols = 6

    # Encabezado superior con fecha y nombres de origen.
    ws.cell(row=1, column=col_nivel).value = datetime.datetime.now().strftime(
        "%m/%d/%y %I:%M %p"
    )
    ws.cell(row=1, column=col_nivel).font = font_bold
    ws.cell(row=1, column=col_nivel).alignment = left

    ws.cell(row=1, column=col_qty_new).value = "Charted_BOM_CapXC"
    ws.cell(row=1, column=col_qty_new).font = font_bold
    ws.cell(row=1, column=col_qty_new).alignment = right

    ws.cell(row=1, column=col_qty_old).value = "MFGPro_CAPH"
    ws.cell(row=1, column=col_qty_old).font = font_bold
    ws.cell(row=1, column=col_qty_old).alignment = center

    titulos_fila2 = {
        col_nivel: "Nivel",
        col_part: "Int.Part Nbr.",
        col_qty_new: "Charted_BOM_CapXC",
        col_qty_old: "MFGPro_CAPH",
        col_unit: "Unit",
        col_desc: "Part Description",
    }

    if not df_result.empty:
        # Si hay datos, el encabezado muestra el modelo/issue del primer bloque.
        primer_issue = df_result.iloc[0].get("_modelo_issue", df_result.iloc[0]["Model"])
        primer_base = df_result.iloc[0]["Model"]
        titulos_fila2[col_qty_new] = primer_issue
        titulos_fila2[col_qty_old] = primer_base

    for col_idx, titulo in titulos_fila2.items():
        cell = ws.cell(row=2, column=col_idx)
        cell.value = titulo
        cell.alignment = left
        cell.font = font_bold

    # Cuerpo del reporte: una fila por diferencia encontrada.
    for i, row_data in df_result.iterrows():
        fila_excel = i + 3
        valores = [
            (col_nivel, row_data["Model"]),
            (col_part, row_data["Part Nbr."]),
            (col_qty_new, row_data["Qty New"]),
            (col_qty_old, row_data["Qty Old"]),
            (col_unit, row_data["UOM"]),
            (col_desc, row_data["Description"]),
        ]

        for col_idx, valor in valores:
            cell = ws.cell(row=fila_excel, column=col_idx)
            cell.value = valor
            cell.font = font_normal
            cell.alignment = right if col_idx in (col_qty_new, col_qty_old) else left

    for col_ltr, width in column_widths.items():
        ws.column_dimensions[col_ltr].width = width

    # Filtro y vista final para que el archivo abra listo para revisar.
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{ws.max_row}"

    ws.sheet_view.zoomScale = 85
    ws.sheet_view.showGridLines = False

    wb.save(ruta_salida)
